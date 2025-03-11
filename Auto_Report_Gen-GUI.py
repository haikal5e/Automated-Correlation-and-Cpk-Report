# Import necessary libraries
import pandas as pd  # For data manipulation
import numpy as np  # For numerical operations
import warnings  # To handle warnings
import sys  # For system-related functions
import os  # For operating system tasks
import platform  # For platform information
import time  # For time-related functions
from openpyxl import Workbook  # For working with Excel files
from openpyxl.utils import get_column_letter  # To convert column numbers to letters
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill  # For styling Excel cells
from datetime import datetime  # For date and time handling
import tkinter as tk
from tkinter import messagebox, simpledialog, filedialog, ttk

# Ignore warnings to keep output clean
warnings.filterwarnings('ignore')

def clear_console():
    """
    Clear the console screen.

    This function checks the operating system and executes the appropriate command
    to clear the console. It uses 'cls' for Windows and 'clear' for Unix/Linux/Mac.

    Returns:
        None
    """
    if platform.system() == "Windows":
        os.system('cls')  # For Windows
    else:
        os.system('clear')  # For Unix/Linux/Mac

def display_banner():
    """
    Display a decorative banner in the console with #a29bfe color.

    This function prints a stylized banner that includes information about the 
    Auto Correlation and Cpk Report Generator, along with the author's name, 
    email, and GitHub link. The banner is displayed in #a29bfe color and is followed 
    by a brief pause.

    Returns:
        None
    """
    # ANSI escape codes for #a29bfe (RGB: 162, 155, 254) and reset
    custom_color = '\033[38;2;162;155;254m'
    reset_color = '\033[0m'
    
    banner = '''
┏┓       ┏┓        ┓   •           ┓  ┏┓  ┓   ┳┓           ┏┓               
┣┫┓┏╋┏┓  ┃ ┏┓┏┓┏┓┏┓┃┏┓╋┓┏┓┏┓  ┏┓┏┓┏┫  ┃ ┏┓┃┏  ┣┫┏┓┏┓┏┓┏┓╋  ┃┓┏┓┏┓┏┓┏┓┏┓╋┏┓┏┓
┛┗┗┻┗┗┛  ┗┛┗┛┛ ┛ ┗ ┗┗┻┗┗┗┛┛┗  ┗┻┛┗┗┻  ┗┛┣┛┛┗  ┛┗┗ ┣┛┗┛┛ ┗  ┗┛┗ ┛┗┗ ┛ ┗┻┗┗┛┛ 
                                        ┛         ┛                                                                                                              
Auto Correlation and Cpk Report Generator
Coded by Mohamad Haikal bin Mohamad Nazari
Email: mohamadhaikal.mohamadnazari@tessolve.com
Github: https://github.com/haikal5e
    '''
    print(f"{custom_color}{banner}{reset_color}")
    time.sleep(1)

def wait_for_enter():
    # Blue text for the prompt
    print('\033[38;2;162;155;254mPlease tap "Enter" to start\033[0m', end='')
    input()  # Wait for Enter press

def thank_you():
    """
    Display a thank you message in the console using #55efc4 color.

    This function prints a stylized thank you message in a mint green color (#55efc4). 
    The message is displayed in a decorative format and is followed by a brief pause.

    Returns:
        None
    """
    # ANSI escape codes for #55efc4 (RGB: 85, 239, 196)
    mint_green = '\033[38;2;85;239;196m'
    reset_color = '\033[0m'
    
    thank = '''
┏┳┓┓     ┓   ┓┏    
 ┃ ┣┓┏┓┏┓┃┏  ┗┫┏┓┓┏
 ┻ ┛┗┗┻┛┗┛┗  ┗┛┗┛┗┻
                                                                                                              
    '''
    
    print(f"{mint_green}{thank}{reset_color}")
    time.sleep(1)

def get_product_information():
    product_info = {
        'Test Card Name': '',
        'Part Name': '',
        'Package': '',
        'Lead Count': '',
        'Description': ''
    }
    
    root = tk.Tk()
    root.title("Product Information Entry")
    current_frame = None

    def create_input_screen():
        nonlocal current_frame
        if current_frame:
            current_frame.destroy()
            
        current_frame = tk.Frame(root)
        current_frame.pack(padx=20, pady=20)
        
        entries = {}
        for row, key in enumerate(product_info):
            tk.Label(current_frame, text=f"{key}:").grid(row=row, column=0, sticky='w', pady=5)
            entry = tk.Entry(current_frame, width=30)
            entry.grid(row=row, column=1, pady=5)
            entries[key] = entry

        def validate_and_continue():
            for key, entry in entries.items():
                if not entry.get().strip():
                    messagebox.showerror("Error", f"{key} cannot be empty")
                    return
            for key in product_info:
                product_info[key] = entries[key].get().strip()
            create_review_screen()

        tk.Button(
            current_frame, 
            text="Submit", 
            command=validate_and_continue
        ).grid(row=len(product_info)+1, columnspan=2, pady=10)

    def create_review_screen():
        nonlocal current_frame
        current_frame.destroy()
        
        current_frame = tk.Frame(root)
        current_frame.pack(padx=20, pady=20)

        for row, (key, value) in enumerate(product_info.items()):
            tk.Label(current_frame, text=f"{key}:", width=15, anchor='w').grid(row=row, column=0, sticky='w')
            tk.Label(current_frame, text=value, width=25, anchor='w').grid(row=row, column=1, sticky='w')
            
            def make_edit_handler(k):
                def edit_handler():
                    get_new_value(k)
                return edit_handler
            
            tk.Button(
                current_frame, 
                text="Edit", 
                command=make_edit_handler(key)
            ).grid(row=row, column=2, padx=5)

        def confirm_and_exit():
            root.destroy()
            
        tk.Button(
            current_frame, 
            text="Confirm Information", 
            command=confirm_and_exit,
            bg='green', fg='white'
        ).grid(row=len(product_info)+1, columnspan=3, pady=15)

    def get_new_value(field):
        dialog = tk.Toplevel(root)
        dialog.title(f"Edit {field}")
        
        tk.Label(dialog, text=f"New {field}:").pack(pady=5)
        new_value_entry = tk.Entry(dialog, width=30)
        new_value_entry.insert(0, product_info[field])
        new_value_entry.pack(pady=5)

        def save_new_value():
            new_value = new_value_entry.get().strip()
            if not new_value:
                messagebox.showerror("Error", "Field cannot be empty")
                return
            product_info[field] = new_value
            dialog.destroy()
            create_review_screen()

        tk.Button(dialog, text="Save", command=save_new_value).pack(pady=10)

    def on_closing():
        if messagebox.askokcancel("Quit", "Do you want to cancel entry?"):
            root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing)
    create_input_screen()
    root.mainloop()
    
    return product_info

def get_setup_information():
    setup_info = {
        'Tester ID': '',
        'Reference Board': '',
        'New Board ID': '',
        'Test Program': ''
    }
    
    root = tk.Tk()
    root.title("Setup Configuration")
    current_frame = None

    def create_input_screen():
        nonlocal current_frame
        if current_frame:
            current_frame.destroy()
            
        current_frame = tk.Frame(root)
        current_frame.pack(padx=20, pady=20)
        
        entries = {}
        for row, key in enumerate(setup_info):
            tk.Label(current_frame, text=f"{key}:").grid(row=row, column=0, sticky='w', pady=5)
            entry = tk.Entry(current_frame, width=30)
            entry.grid(row=row, column=1, pady=5)
            entries[key] = entry

        def validate_and_continue():
            for key, entry in entries.items():
                if not entry.get().strip():
                    messagebox.showerror("Error", f"{key} cannot be empty")
                    return
            for key in setup_info:
                setup_info[key] = entries[key].get().strip()
            create_review_screen()

        tk.Button(
            current_frame, 
            text="Submit", 
            command=validate_and_continue
        ).grid(row=len(setup_info)+1, columnspan=2, pady=10)

    def create_review_screen():
        nonlocal current_frame
        current_frame.destroy()
        
        current_frame = tk.Frame(root)
        current_frame.pack(padx=20, pady=20)

        for row, (key, value) in enumerate(setup_info.items()):
            tk.Label(current_frame, text=f"{key}:", width=15, anchor='w').grid(row=row, column=0, sticky='w')
            tk.Label(current_frame, text=value, width=25, anchor='w').grid(row=row, column=1, sticky='w')
            
            def make_edit_handler(k):
                def edit_handler():
                    get_new_value(k)
                return edit_handler
            
            tk.Button(
                current_frame, 
                text="Edit", 
                command=make_edit_handler(key)
            ).grid(row=row, column=2, padx=5)

        def confirm_and_exit():
            root.destroy()
            
        tk.Button(
            current_frame, 
            text="Confirm Setup", 
            command=confirm_and_exit,
            bg='green', fg='white'
        ).grid(row=len(setup_info)+1, columnspan=3, pady=15)

    def get_new_value(field):
        dialog = tk.Toplevel(root)
        dialog.title(f"Edit {field}")
        
        tk.Label(dialog, text=f"New {field}:").pack(pady=5)
        new_value_entry = tk.Entry(dialog, width=30)
        new_value_entry.insert(0, setup_info[field])
        new_value_entry.pack(pady=5)

        def save_new_value():
            new_value = new_value_entry.get().strip()
            if not new_value:
                messagebox.showerror("Error", "Field cannot be empty")
                return
            setup_info[field] = new_value
            dialog.destroy()
            create_review_screen()

        tk.Button(dialog, text="Save", command=save_new_value).pack(pady=10)

    def on_closing():
        if messagebox.askokcancel("Quit", "Do you want to cancel setup configuration?"):
            root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing)
    create_input_screen()
    root.mainloop()
    
    return setup_info

def get_data():
    root = tk.Tk()
    root.withdraw()

    # Case-insensitive CSV check
    def is_csv_file(path):
        return path and path.lower().endswith('.csv')

    # Get number of boards
    def get_num_boards():
        while True:
            num = simpledialog.askinteger("Input", "Enter the number of New Boards (NB) (1-9):")
            if num is None: return None
            if 1 <= num <= 9: return num
            messagebox.showerror("Error", "Please enter between 1-9")

    # Get number of units
    def get_num_units():
        while True:
            num = simpledialog.askinteger("Input", "Enter the number of units tested (1-9):")
            if num is None: return None
            if 1 <= num <= 9: return num
            messagebox.showerror("Error", "Please enter between 1-9")

    # Main data collection
    num_boards = get_num_boards()
    num_units = get_num_units()

    # Amount confirmation
    while not messagebox.askyesno("Confirmation",
                                  f"Confirm: Total {num_boards} new board(s) (NB) with {num_units} units tested"):
        num_boards = get_num_boards()
        num_units = get_num_units()

    # File collection with case-insensitive checks
    board_files = []
    for board_idx in range(num_boards):
        while True:
            files = filedialog.askopenfilenames(
                title=f"Select {num_units} CSVs for NB{board_idx + 1}",
                filetypes=[("CSV Files", "*.csv;*.CSV;*.Csv")]
            )
            if len(files) != num_units:
                messagebox.showerror("Error", f"Need {num_units} files, got {len(files)}")
                continue
            if not all(is_csv_file(f) for f in files):
                messagebox.showerror("Error", "All files must be CSV format")
                continue
            board_files.append([os.path.basename(f) for f in files])
            break

    rb_files = []
    for unit_idx in range(num_units):
        while True:
            path = filedialog.askopenfilename(
                title=f"Select CSV for RB U{unit_idx + 1}",
                filetypes=[("CSV Files", "*.csv;*.CSV;*.Csv")]
            )
            if not is_csv_file(path):
                messagebox.showerror("Error", "Invalid RB file format")
                continue
            rb_files.append(os.path.basename(path))
            break

    limit_file = ""
    while True:
        path = filedialog.askopenfilename(
            title="Select Limit CSV",
            filetypes=[("CSV Files", "*.csv;*.CSV;*.Csv")]
        )
        if not is_csv_file(path):
            messagebox.showerror("Error", "Invalid limit file format")
            continue
        limit_file = os.path.basename(path)
        break

    # Review window with case-insensitive validation
    def create_review_window():
        review_win = tk.Toplevel()
        review_win.title("File Review")
        review_win.geometry("800x600")

        notebook = ttk.Notebook(review_win)

        # Board files tab
        board_frame = ttk.Frame(notebook)
        board_list = create_scrolled_list(
            board_frame,
            [f"NB{i + 1}: {', '.join(files)}" for i, files in enumerate(board_files)],
            lambda idx: edit_board_files(idx, board_files, board_list)
        )
        notebook.add(board_frame, text="NB files")

        # RB files tab
        rb_frame = ttk.Frame(notebook)
        rb_list = create_scrolled_list(
            rb_frame,
            [f"RB U{i + 1}: {path}" for i, path in enumerate(rb_files)],
            lambda idx: edit_rb_file(idx, rb_files, rb_list)
        )
        notebook.add(rb_frame, text="RB files")

        # Limit file tab
        limit_frame = ttk.Frame(notebook)
        limit_label = ttk.Label(limit_frame, text=f"Current Limit File:\n{limit_file}")
        limit_label.pack(pady=10)
        ttk.Button(limit_frame, text="Change Limit File",
                   command=lambda: update_limit_file(limit_frame, limit_label)).pack()
        notebook.add(limit_frame, text="Limit file")

        notebook.pack(expand=True, fill=tk.BOTH)
        ttk.Button(review_win, text="CONFIRM ALL",
                   command=review_win.destroy).pack(pady=20)

        review_win.grab_set()
        review_win.wait_window()

    def create_scrolled_list(parent, items, edit_callback):
        frame = ttk.Frame(parent)
        listbox = tk.Listbox(frame, selectmode=tk.SINGLE, width=90)
        scroll = ttk.Scrollbar(frame, command=listbox.yview)

        for item in items:
            listbox.insert(tk.END, item)

        listbox.config(yscrollcommand=scroll.set)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        ttk.Button(frame, text="Edit Selected",
                   command=lambda: handle_edit(listbox, edit_callback)).pack(pady=5)
        frame.pack(expand=True, fill=tk.BOTH)
        return listbox

    def handle_edit(listbox, callback):
        if selection := listbox.curselection():
            callback(selection[0])

    def edit_board_files(index, files, list_widget):
        new_files = filedialog.askopenfilenames(
            title=f"Select {num_units} files for NB{index + 1}",
            filetypes=[("CSV Files", "*.csv;*.CSV;*.Csv")]
        )
        if len(new_files) != num_units:
            messagebox.showerror("Error", f"Need exactly {num_units} files")
            return
        if not all(is_csv_file(f) for f in new_files):
            messagebox.showerror("Error", "All files must be CSV format")
            return

        files[index] = [os.path.basename(f) for f in new_files]
        list_widget.delete(index)
        list_widget.insert(index, f"NB{index + 1}: {', '.join(files[index])}")
        list_widget.selection_set(index)

    def edit_rb_file(index, files, list_widget):
        new_path = filedialog.askopenfilename(
            title=f"Select new file for RB U{index + 1}",
            filetypes=[("CSV Files", "*.csv;*.CSV;*.Csv")]
        )
        if not is_csv_file(new_path):
            messagebox.showerror("Error", "Invalid file format")
            return

        files[index] = os.path.basename(new_path)
        list_widget.delete(index)
        list_widget.insert(index, f"RB U{index + 1}: {files[index]}")
        list_widget.selection_set(index)

    def update_limit_file(parent, label):
        nonlocal limit_file
        new_path = filedialog.askopenfilename(
            title="Select new limit file",
            filetypes=[("CSV Files", "*.csv;*.CSV;*.Csv")]
        )
        if not is_csv_file(new_path):
            messagebox.showerror("Error", "Invalid file format")
            return

        limit_file = os.path.basename(new_path)
        label.config(text=f"Current Limit File:\n{limit_file}")
        parent.update()

    create_review_window()
    root.destroy()

    return board_files, rb_files, num_boards, num_units, limit_file

def process_dataframes(dataframes):
    """
    Process a list of DataFrames with the following steps:
    - Remove rows with NaN values
    - Drop the 'Test #' column
    - Set 'Description' as the index
    - Transpose the DataFrame
    - Remove the 'Units' index
    - Reset the index
    - Convert all data to float

    Parameters:
    - dataframes: List of DataFrames to process

    Returns:
    - List of processed DataFrames
    """
    processed_dataframes = []

    for df in dataframes:        
        df = df.dropna(thresh=df.shape[1] - 1)  # Remove rows with too many NaNs
        df = df.set_index('Description')  # Use 'Description' as the index
        df = df.drop(df.columns[0], axis=1)  # Drop the first column ('Test #')
        df = df.T  # Transpose the DataFrame
        df = df.drop(index='Units', errors='ignore')  # Remove 'Units' index if it exists
        df = df.reset_index(drop=True)  # Reset the index
        df = df.astype(float)  # Convert all values to float
        
        processed_dataframes.append(df)  # Add the processed DataFrame to the list

    return processed_dataframes

def mean_shift(row, index):
    """
    Calculate the mean shift percentage from a given row and index.

    This function computes the mean shift percentage using the 'Delta Mean' value 
    and the corresponding low and high limits from the `limit` DataFrame. It handles 
    cases where limits are missing or 'Delta Mean' is zero.

    Parameters:
    - row (pandas.Series): A row containing 'Delta Mean'.
    - index (int): The index for retrieving limits from the `limit` DataFrame.
    
    Returns:
    - float: Mean shift percentage rounded to five decimal places. Returns NaN 
      if limits are invalid or NaN, and returns 0 if 'Delta Mean' is zero.

    Calculation:
    - Mean shift percentage = (Delta Mean) / (High Limit - Low Limit) * 100
    """
    row2 = limit.iloc[index]
    
    low_limit = float(row2[2]) if len(row2) > 2 else np.nan
    high_limit = float(row2[3]) if len(row2) > 3 else np.nan

    if np.isnan(low_limit) or np.isnan(high_limit):
        return np.nan
    
    if row.get('Delta Mean', 0) == 0:
        return 0
    
    return np.round(row['Delta Mean'] / (high_limit - low_limit) * 100, 5)

def mean_shift_crit(row, index):
    """
    Evaluate the mean shift criteria based on the provided row data.

    This function checks the 'Mean Shift' and 'Delta Mean' values against a specified 
    standard deviation limit. It returns "Passed", "Failed", or "For check" based on 
    the evaluation criteria.

    Parameters:
    - row (pandas.Series): A row of data, typically from a DataFrame, which should contain:
        - 'Mean Shift': The mean shift value to evaluate.
        - 'Delta Mean': The delta mean value.
    - index (int): The index to retrieve the corresponding row from the `limit` DataFrame,
      which should contain the standard deviation limit at column index 4.
    
    Returns:
    - str: "Passed", "Failed", or "For check" based on the evaluation criteria:
        - "Passed": If the 'Mean Shift' is less than or equal to 5, or if 'Mean Shift' 
          is NaN and 'Delta Mean' is less than or equal to the standard deviation limit.
        - "Failed": If the 'Mean Shift' is greater than 5.
        - "For check": If 'Mean Shift' is NaN and 'Delta Mean' exceeds the standard 
          deviation limit.

    Notes:
    - The standard deviation limit is retrieved from the `limit` DataFrame at the given index.
    - The function assumes that the relevant columns in the `row` Series and the `limit` 
      DataFrame are correctly populated.
    """
    row2 = limit.iloc[index]
    
    mean_shf = row['Mean Shift']
    d_mean = row['Delta Mean']
    sdlot = row2[4]

    # Check the conditions
    if np.isnan(mean_shf):  # Check if mean_shf is NaN
        return "Passed" if d_mean <= sdlot else "For check"
    else:
        return "Passed" if mean_shf <= 5 else "Failed"

def sd_ratio(row):
    """
    Calculate the standard deviation ratio based on the provided row data.

    This function computes the ratio of two values from the row. If either of the 
    values is zero, the function returns 0. Otherwise, it returns the ratio rounded 
    to six decimal places.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - Column index 7: The standard deviation for RB at certain unit.
        - Column index 9: The standard deviation for certain NB at certain unit.

    Returns:
    - float: The standard deviation ratio rounded to six decimal places, or 0 if 
      either value is zero.
    """
    if row[1] == 0:
        return 0
    elif row[3] == 0:
        return 0
    else:
        return np.round(row[3] / row[1], 6)

def sd_ratio_crit(row):
    """
    Evaluate the standard deviation ratio criteria based on the provided row data.

    This function checks the 'SD Ratio' value against a threshold of 1.5. 
    It returns "Passed" if the ratio is less than or equal to 1.5, 
    and "For check" otherwise.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - 'SD Ratio': The standard deviation ratio to evaluate.

    Returns:
    - str: "Passed" if the SD Ratio is less than or equal to 1.5, 
           or "For check" if it exceeds 1.5.
    """
    if row["SD Ratio"] <= 1.5:
        return "Passed"
    else:
        return "For check"

def eva_status(row):
    """
    Evaluate the overall status based on mean shift and standard deviation ratio criteria.

    This function checks the evaluation status of 'Mean Shift Criteria' and 
    'SD Ratio Criteria'. It returns "Passed" if both criteria are passed, 
    "Failed" if the mean shift criteria has failed, and "For check" 
    for any other combination of statuses.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - 'Mean Shift Criteria': The evaluation result of the mean shift.
        - 'SD Ratio Criteria': The evaluation result of the standard deviation ratio.

    Returns:
    - str: "Passed" if both criteria are passed, 
           "Failed" if the mean shift criteria has failed, 
           or "For check" for any other combination.
    """
    if row['Mean Shift Criteria'] == "Passed" and row['SD Ratio Criteria'] == "Passed":
        return "Passed"
    elif row['Mean Shift Criteria'] == "Failed":
        return "Failed"
    else:
        return "For check"

def calculate_cp_rb(row):
    """
    Calculate the Cp and RB value based on the provided row data.

    This function computes the Cp and Rb value using the formula:
    (upper_spec_limit - lower_spec_limit) / (6 * std_dev). It checks for specific conditions 
    before performing the calculation, returning NaN if any of the 
    required values are zero or missing.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - row[2]: The lower specification limit.
        - row[3]: The upper specification limit.
        - row[7]: The standard deviation for RB

    Returns:
    - float: The calculated Cp for RB and the value rounded to two decimal places, 
             or NaN if any of the required values are zero or missing.
    """
    # Check if row[7] is 0
    if row[7] == 0:
        return np.nan
    
    # Check for empty or "NA" values
    if row[2] is np.nan or row[3] is np.nan or row[7] is np.nan:
        return np.nan
    
    # Perform the calculation
    result = (row[3] - row[2]) / (6 * row[7])
    
    # Round the result to 2 decimal places
    return round(result, 2)

def calculate_cpk_rb(row):
    """
    Calculate the CpK and RB value based on the provided row data.

    This function computes the CpK value using the available specification limits 
    and standard deviation. It checks for NaN values and ensures that the standard 
    deviation is not zero before performing the calculation. The function returns 
    the CpK value rounded to two decimal places along with a decision message 
    indicating the capability status.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - row[2]: The lower specification limit.
        - row[3]: The upper specification limit.
        - row[6]: The mean for RB at certain unit.
        - row[7]: The standard deviation for RB

    Returns:
    - tuple: A tuple containing:
        - float: The calculated CpK value rounded to two decimal places, 
                 or NaN if the required values are missing or invalid.
        - str: A decision message indicating the capability status ("Not capable" 
               or "Good capable").
    """
    # Check for NaN values and if row[7] is 0
    if np.isnan(row[2]) and np.isnan(row[3]) or row[7] == 0:
        return np.nan, "N/A"
    
    # Calculate cpk based on the available values
    if np.isnan(row[2]):
        cpk = (row[3] - row[6]) / (3 * row[7])
    elif np.isnan(row[3]):
        cpk = (row[6] - row[2]) / (3 * row[7])
    else:
        cpk = min(row[3] - row[6], row[6] - row[2]) / (3 * row[7])

    # Round cpk to 2 decimal places
    cpk = round(cpk, 2)

    # Concise decision statements
    if cpk < 1.3:
        decision_message = "Not capable"
    else:
        decision_message = "Good capable"

    return cpk, decision_message

def calculate_cp_nb(row):
    """
    Calculate the Cp value based on the provided row data.

    This function computes the Cp value using the formula:
    (upper_spec_limit - lower_spec_limit) / (6 * std_dev). It checks for specific conditions 
    before performing the calculation, returning NaN if any of the 
    required values are zero or missing.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - row[2]: The lower specification limit.
        - row[3]: The upper specification limit.
        - row[12]: The standard deviation for certain NB

    Returns:
    - float: The calculated Cp value rounded to two decimal places, 
             or NaN if any of the required values are zero or missing.
    """
    # Check if row[12] is 0
    if row[12] == 0:
        return np.nan
    
    # Check for empty or "NA" values
    if np.isnan(row[2]) or np.isnan(row[3]) or np.isnan(row[12]):
        return np.nan
    
    # Perform the calculation
    result = (row[3] - row[2]) / (6 * row[12])
    
    # Round the result to 2 decimal places
    return round(result, 2)

def calculate_cpk_nb(row):
    """
    Calculate the CpK value based on the provided row data.

    This function computes the CpK value using the available specification limits 
    and standard deviation. It checks for NaN values and ensures that the standard 
    deviation is not zero before performing the calculation. The function returns 
    the CpK value rounded to two decimal places along with a decision message 
    indicating the capability status.

    Parameters:
    - row: A pandas Series representing a row of data, which should contain:
        - row[2]: The lower specification limit.
        - row[3]: The upper specification limit.
        - row[11]: The mean for certain NB.
        - row[12]: The standard deviation for certain NB.

    Returns:
    - tuple: A tuple containing:
        - float: The calculated CpK value rounded to two decimal places, 
                 or NaN if the required values are missing or invalid.
        - str: A decision message indicating the capability status ("Not capable" 
               or "Good capable").
    """
    # Check for NaN values and if row[12] is 0
    if np.isnan(row[2]) and np.isnan(row[3]) or row[12] == 0:
        return np.nan, "N/A"
    
    # Calculate cpk based on the available values
    if np.isnan(row[2]):
        cpk = (row[3] - row[11]) / (3 * row[12])
    elif np.isnan(row[3]):
        cpk = (row[11] - row[2]) / (3 * row[12])
    else:
        cpk = min(row[3] - row[11], row[11] - row[2]) / (3 * row[12])

    # Round cpk to 2 decimal places
    cpk = round(cpk, 2)

    # Concise decision statements
    if cpk < 1.3:
        decision_message = "Not capable"
    else:
        decision_message = "Good capable"

    return cpk, decision_message

def check_value(col):
    """
    Check the value in the specified column and return a corresponding message.

    This function evaluates the value in the fourth element of the provided 
    column (index 3) which is Failed test all units and returns a message based on its content. 

    Parameters:
    - col: A list or array-like structure where the fourth element (index 3) 
           is evaluated in this "Failed test all units" column.

    Returns:
    - str: A message indicating the status based on the value:
        - An empty string if the value is an empty string.
        - "Good to release if no concern" if the value is 0.
        - "Not acceptable" for any other value.
    """
    if col[3] == "":
        return ""
    elif col[3] == 0:
        return "Good to release if no concern"
    else:
        return "Not acceptable"

def autosize_columns(worksheet):
    """
    Auto-adjust the width of columns in the given worksheet.

    This function iterates through all rows in the specified worksheet and calculates 
    the maximum length of the content in each column. It then sets the width of each 
    column based on the calculated maximum lengths, adding a small padding for better 
    visibility.

    Parameters:
    - worksheet: An instance of an openpyxl worksheet where the column widths 
                  need to be adjusted.
    
    Returns:
    - None: This function modifies the worksheet in place and does not return a value.
    """
    column_widths = []

    # Iterate through all rows in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            if cell is not None:
                cell_length = len(str(cell))  # Get the length of the cell content
                if len(column_widths) > i:
                    if cell_length > column_widths[i]:
                        column_widths[i] = cell_length
                else:
                    column_widths.append(cell_length)

    # Set the width of each column based on the calculated maximum lengths
    for i, column_width in enumerate(column_widths, 1):  # Start at 1 for column indexing
        worksheet.column_dimensions[get_column_letter(i)].width = column_width + 0.5  # Adding padding

# Display the banner to the user
display_banner()

wait_for_enter()

# Retrieve product information
product_info = get_product_information()

# Retrieve setup information
setup_info = get_setup_information()

# Obtain data related to files, number of boards, number of units, and limits
nb_file, rb_file, num_boards, num_units, limit = get_data()

# Read the CSV file into a DataFrame
limit = pd.read_csv(limit)

# Reset the index of the DataFrame, dropping the old index
limit.reset_index(drop=True, inplace=True)

# Convert the 'Test #' column to strings and remove '\t'
limit.iloc[:, 0] = limit.iloc[:, 0].astype(str).str.replace(r'\t', '', regex=True)

# Remove rows with too many NaNs
limit.dropna(thresh=limit.shape[1] - 3, inplace=True)  

# Fill any NaN values with numpy's NaN
limit.fillna(np.nan, inplace=True)

# Convert the values in the third column to float
limit.iloc[:, 2] = limit.iloc[:, 2].astype('float')

# Convert the values in the fourth column to float
limit.iloc[:, 3] = limit.iloc[:, 3].astype('float')

# Initialize an empty list to hold the DataFrames
rb = []

# Loop through the unit numbers
for i in range(num_units):
    # Construct the filename based on the unit number
    file_path = rb_file[i]
    
    # Read the CSV file and append the DataFrame to the list
    try:
        df = pd.read_csv(file_path)  # Read the CSV file into a DataFrame
        rb.append(df)  # Append the DataFrame to the list
    except FileNotFoundError:
        print(f"File not found: {file_path}")  # Handle the case where the file is not found
    except Exception as e:
        print(f"An error occurred while reading {file_path}: {e}")  # Handle any other exceptions

# Initialize an empty list to hold the DataFrames for each board
nb = []

# Loop to get input paths for each board and its units
for i in range(num_boards):
    board_units = []  # Initialize a list to hold the DataFrames for the current board
    
    for j in range(num_units):
        # Construct the input path based on board and unit numbers
        file_path = nb_file[i][j]

        # Read the DataFrame from the constructed input path
        try:
            df = pd.read_csv(file_path)  # Read the CSV file into a DataFrame
            board_units.append(df)  # Append the DataFrame for the current unit
            
        except FileNotFoundError:
            print(f"File not found: {file_path}")  # Handle the case where the file is not found
        except Exception as e:
            print(f"An error occurred while reading {file_path}: {e}")  # Handle any other exceptions
    
    # Append the current board's units to the main list
    nb.append(board_units)  # Add the list of DataFrames for the current board to the main list

# Process the list of DataFrames for the 'rb' variable
rb_mod = process_dataframes(rb)

# Initialize an empty list to hold processed DataFrames for each board
nb_mod = []

# Loop through each board in the 'nb' list
for board in nb:
    # Process the DataFrames for the current board
    processed_units = process_dataframes(board)
    
    # Append processed DataFrames for each board to the main list
    nb_mod.append(processed_units)  # Add the processed units for the current board

# Initialize empty lists to hold the mean and standard deviation DataFrames
mean_rb = []
std_rb = []

# Loop through each modified DataFrame in the 'rb_mod' list
for df_mod in rb_mod:
    # Calculate the mean for each column and format it to six decimal places
    mean = df_mod.mean().apply(lambda x: f'{x:.6f}').astype(float)
    
    # Calculate the standard deviation for each column (using population standard deviation) and format it to five decimal places
    std = df_mod.std(ddof=0).apply(lambda x: f'{x:.5f}').astype(float)
    
    # Append the mean and standard deviation to their respective lists
    mean_rb.append(mean)
    std_rb.append(std)

# Reset the index of each mean DataFrame to create a clean list of DataFrames
mean_rb_clean = [df.reset_index(drop=True) for df in mean_rb]

# Reset the index of each standard deviation DataFrame to create a clean list of DataFrames
std_rb_clean = [df.reset_index(drop=True) for df in std_rb]

# Create a list of column names for the mean DataFrames
columns_mean_rb = [f'Mean RB U{i+1}' for i in range(0, num_units)]

# Create a list of column names for the standard deviation DataFrames
columns_std_rb = [f'SD RB U{i+1}' for i in range(0, num_units)]

# Initialize empty DataFrames to hold the real mean and standard deviation values
realrbmean = pd.DataFrame()
realrbstd = pd.DataFrame()

# Loop through each column name and corresponding mean value
for col, val in zip(columns_mean_rb, mean_rb_clean):
    realrbmean[col] = val  # Assign the mean values to the corresponding column in the DataFrame

# Loop through each column name and corresponding standard deviation value
for col, val in zip(columns_std_rb, std_rb_clean):
    realrbstd[col] = val  # Assign the standard deviation values to the corresponding column in the DataFrame

# Assuming nb_mod is a list of lists of NumPy arrays instead of DataFrames
mean_nb_clean = []  # Initialize the mean list outside the loop
std_nb_clean = []   # Initialize the standard deviation list outside the loop

# Iterate through each inner list in nb_mod
for inner_list in nb_mod:
    mean_row = []  # Temporary list to hold means for the current inner list
    std_row = []   # Temporary list to hold standard deviations for the current inner list
    
    # Iterate through each NumPy array in the inner list
    for arr_mod in inner_list:
        # Calculate mean and standard deviation using NumPy
        mean = np.mean(arr_mod, axis=0)  # Calculate mean along the specified axis
        std = np.std(arr_mod, axis=0, ddof=0)  # Calculate standard deviation along the specified axis
        
        # Append results to respective temporary lists
        mean_row.append(np.round(mean, 6).tolist())  # Round the mean and convert to list
        std_row.append(np.round(std, 5).tolist())    # Round the standard deviation and convert to list
    
    # Append the temporary lists to the main 2D lists
    mean_nb_clean.append(mean_row)  # Add the mean row to the main list
    std_nb_clean.append(std_row)     # Add the std row to the main list

# mean_nb_clean and std_nb_clean are now 2D lists of means and standard deviations

# Create 2D lists to hold column names for means and standard deviations
columns_mean_nb = []
columns_std_nb = []

# Fill the lists with column names for each board
for i in range(num_boards + 1):
    # Generate mean column names for the current board
    mean_columns = [f'Mean NB{i+1} U{j+1}' for j in range(0, num_units)]
    
    # Generate standard deviation column names for the current board
    std_columns = [f'SD NB{i+1} U{j+1}' for j in range(0, num_units)]
    
    # Append the mean columns for the current board to the list
    columns_mean_nb.append(mean_columns)
    
    # Append the standard deviation columns for the current board to the list
    columns_std_nb.append(std_columns)

# Create a list to hold the empty DataFrames for means
realnbmean = []

# Create a list to hold the empty DataFrames for standard deviations
realnbstd = []

# Loop to create empty DataFrames and append them to the realnbmean list
for _ in range(num_boards):
    realnbmean.append(pd.DataFrame())  # Append an empty DataFrame for each board

# Loop to create empty DataFrames and append them to the realnbstd list
for _ in range(num_boards):
    realnbstd.append(pd.DataFrame())  # Append an empty DataFrame for each board

# Populate the realnbmean DataFrames with mean values
for i in range(num_boards):
    for col, val in zip(columns_mean_nb[i], mean_nb_clean[i]):
        realnbmean[i][col] = val  # Assign mean values to the corresponding columns

# Populate the realnbstd DataFrames with standard deviation values
for i in range(num_boards):
    for col, val in zip(columns_std_nb[i], std_nb_clean[i]):
        realnbstd[i][col] = val  # Assign standard deviation values to the corresponding columns

# Initialize a list to hold the calculated limits for each unit
rb_lim_calc = []  # i = unit

# Loop through each unit to calculate limits
for i in range(num_units):
    # Concatenate the limit DataFrame with the mean and standard deviation for the current unit
    temp = pd.concat([limit, realrbmean.iloc[:, i], realrbstd.iloc[:, i]], axis=1)
    
    # Append the concatenated DataFrame to the rb_lim_calc list
    rb_lim_calc.append(temp)

# Create a 2D list to hold the concatenated DataFrames for each board and unit
rb_nbu = []

# Loop through the boards to create the 2D list
for i in range(num_boards):
    temp_list = []  # Temporary list to hold DataFrames for the current board
    for j in range(num_units):
        # Concatenate the DataFrames for the current unit, including limits, means, and standard deviations
        temp = pd.concat([realrbmean.iloc[:,j],realrbstd.iloc[:,j],realnbmean[i].iloc[:, j], realnbstd[i].iloc[:, j]], axis=1)
        
        # Append the concatenated DataFrame to the temporary list
        temp_list.append(temp)
    
    # Append the temporary list of DataFrames for the current board to the main 2D list
    rb_nbu.append(temp_list)

# Loop through each board
for i in range(num_boards):
    # Loop through each unit within the current board
    for j in range(num_units):
        # Calculate the absolute difference between two columns and round it to 6 decimal places
        rb_nbu[i][j]['Delta Mean'] = rb_nbu[i][j].apply(lambda row: np.round(abs(row[2] - row[0]), 6), axis=1)
        
        # Apply the mean_shift function to calculate the mean shift for each row
        # rb_nbu[i][j]['Mean Shift'] = rb_nbu[i][j].apply(mean_shift, axis=1)
        rb_nbu[i][j]['Mean Shift'] = rb_nbu[i][j].apply(lambda row: mean_shift(row, row.name), axis=1)
        
        # Apply the mean_shift_crit function to evaluate the mean shift criteria for each row
        # rb_nbu[i][j]['Mean Shift Criteria'] = rb_nbu[i][j].apply(mean_shift_crit, axis=1)
        rb_nbu[i][j]['Mean Shift Criteria'] = rb_nbu[i][j].apply(lambda row: mean_shift_crit(row, row.name), axis=1)
        
        # Apply the sd_ratio function to calculate the standard deviation ratio for each row
        rb_nbu[i][j]['SD Ratio'] = rb_nbu[i][j].apply(sd_ratio, axis=1)
        
        # Apply the sd_ratio_crit function to evaluate the standard deviation ratio criteria for each row
        rb_nbu[i][j]['SD Ratio Criteria'] = rb_nbu[i][j].apply(sd_ratio_crit, axis=1)
        
        # Apply the eva_status function to determine the result for each unit based on the evaluations
        rb_nbu[i][j][f'Result Unit {j+1}'] = rb_nbu[i][j].apply(eva_status, axis=1)

# Create an empty list to hold the DataFrames for each board
nb_results = [pd.DataFrame() for _ in range(num_boards)]

# Loop through each board to populate the results
for i in range(num_boards):
    # Create a new column for each unit's result within the current board
    for j in range(num_units):
        # Assign the 'Result Unit' data from rb_nbu to the new DataFrame for the current board and unit
        nb_results[i][f'Result NB{i+1} U{j+1}'] = rb_nbu[i][j][f'Result Unit {j+1}']

    # Check if all results for the board are "Passed" and create a new column for the overall result
    all_passed = nb_results[i].eq("Passed").all(axis=1)
    
    # Assign the overall result based on whether all units passed or not
    nb_results[i][f"NB{i+1} Result"] = all_passed.replace({True: 'Passed', False: 'For check'})

nbtrueresult = []

for i in range(num_boards):
    # Debug: Check column count
    #print(f"Columns in nb_results[{i}]: {nb_results[i].shape[1]}")
    
    # Use the last column instead of hardcoding index 4
    board_data = nb_results[i].iloc[:, -1].reset_index(drop=True).rename(f"NB{i+1} Result")
    
    temp = pd.concat([
        limit.reset_index(drop=True).rename(columns=lambda x: f"{x}"),
        board_data
    ], axis=1)

    temp_units = []
    for j in range(num_units):
        # Ensure unit data has unique column names
        unit_df = rb_nbu[i][j].rename(columns=lambda x: f"{x}")
        temp_units.append(unit_df.reset_index(drop=True))

    if temp_units:
        combined_units = pd.concat(temp_units, axis=1)
        final_result = pd.concat([temp, combined_units], axis=1)
    else:
        final_result = temp

    nbtrueresult.append(final_result)

# Concatenate all DataFrames in rb_mod for each unit into a single DataFrame, ignoring the index
rbdf = pd.concat([rb_mod[i] for i in range(num_units)], ignore_index=True)

# Initialize an empty list to hold the concatenated DataFrames for each board
nbdf = []

# Loop through each board to concatenate the results for all units
for i in range(0, num_boards):
    # Concatenate all DataFrames in nb_mod for the current board across all units, ignoring the index
    temp = pd.concat([nb_mod[i][j] for j in range(num_units)], ignore_index=True)
    
    # Append the concatenated DataFrame for the current board to the nbdf list
    nbdf.append(temp)

# Calculate the mean of the rbdf DataFrame, format it to four decimal places, and convert it to float
meanrbdf = rbdf.mean().apply(lambda x: f'{x:.4f}').astype(float).reset_index(drop=True)

# Initialize an empty list to hold the mean results for each board
meannbdf = []

# Loop through each board to calculate the mean of the corresponding DataFrame
for i in range(0, num_boards):
    # Calculate the mean of the current board's DataFrame, format it to four decimal places, and convert it to float
    temp = nbdf[i].mean().apply(lambda x: f'{x:.4f}').astype(float)
    
    # Append the mean result for the current board to the meannbdf list
    meannbdf.append(temp)

# Reset the index for each DataFrame in the meannbdf list
meannbdf = [df.reset_index(drop=True) for df in meannbdf]

# Calculate the standard deviation of the rbdf DataFrame, format it to four decimal places, and convert it to float
stdrbdf = rbdf.std(ddof=0).apply(lambda x: f'{x:.4f}').astype(float).reset_index(drop=True)

# Initialize an empty list to hold the standard deviation results for each board
stdnbdf = []

# Loop through each board to calculate the standard deviation of the corresponding DataFrame
for i in range(0, num_boards):
    # Calculate the standard deviation of the current board's DataFrame, format it to four decimal places, and convert it to float
    temp = nbdf[i].std(ddof=0).apply(lambda x: f'{x:.4f}').astype(float)
    
    # Append the standard deviation result for the current board to the stdnbdf list
    stdnbdf.append(temp)

# Reset the index for each DataFrame in the stdnbdf list
stdnbdf = [df.reset_index(drop=True) for df in stdnbdf]

# Concatenate the mean and standard deviation DataFrames for rbdf along the columns, ignoring the index
meanstdrbdf = pd.concat([meanrbdf, stdrbdf], axis=1, ignore_index=True)

# Set the column names for the concatenated DataFrame
meanstdrbdf.columns = ['Mean RB', 'SD RB']

# Initialize a list to hold the concatenated mean and standard deviation DataFrames for each board
meanstdnbdf = [pd.DataFrame() for _ in range(num_boards)]

# Loop through each board to concatenate the mean and standard deviation DataFrames
for i in range(0, num_boards):
    # Concatenate the mean and standard deviation DataFrames for the current board along the columns, ignoring the index
    meanstdnbdf[i] = pd.concat([meannbdf[i], stdnbdf[i]], axis=1, ignore_index=True)
    
    # Set the column names for the concatenated DataFrame of the current board
    meanstdnbdf[i].columns = [f'Mean NB{i+1}', f'SD NB{i+1}']

# Concatenate the limit DataFrame with the mean and standard deviation DataFrame for rbdf along the columns
rbcpcpk = pd.concat([limit, meanstdrbdf], axis=1)

# Apply the calculate_cp_rb function to each row of the concatenated DataFrame to calculate the Cp value for rb
rbcpcpk["Cp RB"] = rbcpcpk.apply(calculate_cp_rb, axis=1)

# Apply the calculate_cpk_rb function to each row of the concatenated DataFrame to calculate the Cpk values
# The result is expanded into two new columns: 'Cpk RB' and 'Cpk RB Result'
rbcpcpk[['Cpk RB', 'Cpk RB Result']] = rbcpcpk.apply(calculate_cpk_rb, axis=1, result_type='expand')

# Initialize an empty list to hold the DataFrames for each board's Cp and Cpk results
nbcpkresult = []

# Loop through each board to concatenate the rbcpcpk DataFrame with the corresponding mean and standard deviation DataFrame
for i in range(0, num_boards):
    # Concatenate the rbcpcpk DataFrame with the mean and standard deviation DataFrame for the current board
    temp = pd.concat([rbcpcpk, meanstdnbdf[i]], axis=1, ignore_index=False)
    
    # Append the concatenated DataFrame to the nbcpkresult list
    nbcpkresult.append(temp)

# Loop through each board to calculate the Cp values and add them to the corresponding DataFrame
for i in range(0, num_boards):
    # Apply the calculate_cp_nb function to each row of the current board's DataFrame to calculate the Cp value
    nbcpkresult[i][f"Cp NB{i+1}"] = nbcpkresult[i].apply(calculate_cp_nb, axis=1)

# Loop through each board to calculate the Cpk values and add them to the corresponding DataFrame
for i in range(0, num_boards):
    # Apply the calculate_cpk_nb function to each row of the current board's DataFrame to calculate the Cpk values
    # The result is expanded into two new columns: 'Cpk NB' and 'Cpk NB Result'
    nbcpkresult[i][[f'Cpk NB{i+1}', f'Cpk NB{i+1} Result']] = nbcpkresult[i].apply(calculate_cpk_nb, axis=1, result_type='expand')

# Initialize an empty DataFrame to hold the correlation table
corrtable = pd.DataFrame()

# Creating corrtable and adding necessary columns
# Add a column for the test card names, formatted with the product info and board index
corrtable['Test Card'] = [f"{product_info['Test Card Name']}_NB{i+1}" for i in range(num_boards)]

# Add a column for the count of passed tests across all units for each board
corrtable['Passed test all units'] = [nb_results[i][f"NB{i+1} Result"].str.contains('Passed').sum() for i in range(num_boards)]

# Add a column for the count of tests marked 'For check' across all units for each board
corrtable['For Check test all units'] = [nb_results[i][f"NB{i+1} Result"].str.contains('For check').sum() for i in range(num_boards)]

# Add a column for the count of failed tests across all units for each board
corrtable['Failed test all units'] = [nb_results[i][f"NB{i+1} Result"].str.contains('Failed').sum() for i in range(num_boards)]

# Adding unit-specific columns for each unit
for j in range(num_units):
    # Add a column for the count of passed tests for the current unit across all boards
    corrtable[f'Passed test U{j+1}'] = [nb_results[i][f"Result NB{i+1} U{j+1}"].str.contains('Passed').sum() for i in range(num_boards)]
    
    # Add a column for the count of tests marked 'For check' for the current unit across all boards
    corrtable[f'For Check test U{j+1}'] = [nb_results[i][f"Result NB{i+1} U{j+1}"].str.contains('For check').sum() for i in range(num_boards)]
    
    # Add a column for the count of failed tests for the current unit across all boards
    corrtable[f'Failed test U{j+1}'] = [nb_results[i][f"Result NB{i+1} U{j+1}"].str.contains('Failed').sum() for i in range(num_boards)]

# Add a column for the total number of tests conducted (assuming all boards have the same number of results)
corrtable['Total test'] = len(nb_results[0])

# Apply a function to check values and add remarks to the DataFrame
corrtable['Remarks'] = corrtable.apply(check_value, axis=1)

# Set the 'Test Card' column as the index of the DataFrame
corrtable.set_index('Test Card', inplace=True)

# Transpose the DataFrame to switch rows and columns
corrtable = corrtable.T

# Optional: Rename the axes for clarity
corrtable.rename_axis("Test Card", axis=0, inplace=True)
corrtable.rename_axis("Index", axis=1, inplace=True)

# Reset the index to convert the index back into a column
corrtable.reset_index(inplace=True)

# Create a DataFrame from the product_info dictionary, using the index as the first column and 'Details' as the second column
df_product_info = pd.DataFrame.from_dict(product_info, orient='index', columns=['Details']).reset_index()

# Create a DataFrame from the setup_info dictionary, using the index as the first column and 'Details' as the second column
df_setup_info = pd.DataFrame.from_dict(setup_info, orient='index', columns=['Details']).reset_index()

# Rename the columns of the product info DataFrame for clarity
df_product_info.columns = ['Product Info', 'Details']

# Rename the columns of the setup info DataFrame for clarity
df_setup_info.columns = ['Setup Info', 'Details']

# Define the output file name for the Excel report, incorporating the test card name and the current date
current_date = datetime.now().strftime("%d-%m-%Y")  # Get the current date in DD-MM-YYYY format

output_file = f'{product_info["Test Card Name"]}_Correlation_Report_{current_date}.xlsx'

# Create an Excel writer object to write multiple DataFrames to an Excel file
with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
    
    """
    Generates an Excel report containing various data summaries and results.

    This script creates an Excel file with multiple sheets, each containing different 
    DataFrames related to product information, setup information, correlation summaries, 
    and results from various analyses. The report is formatted for clarity, with 
    conditional formatting applied to highlight results based on specific criteria.

    The following sheets are created:
    - 'Info': Contains product information.
    - 'Setup Info': Contains setup information, written below the product info.
    - 'Correlation Summary': Contains a summary of correlation results.
    - 'NB# Correlation Results': Contains results for each NB analysis, with conditional formatting.
    - 'NB# CPK': Contains CPK results, also with conditional formatting.

    The report includes:
    - Dynamic column width adjustments based on content.
    - Autofilters for easy data navigation.
    - Conditional formatting to visually distinguish between different result statuses.
    - Borders and fill colors for improved readability.

    The output file is named using the test card name and the current date.
    """
    
    # Write additional DataFrames to specific sheets first
    df_product_info.to_excel(writer, sheet_name='Info', index=False)  # Write product info to 'Info' sheet
    df_setup_info.to_excel(writer, sheet_name='Info', index=False, startrow=len(df_product_info) + 2, header=True)  # Write setup info below product info
    corrtable.to_excel(writer, sheet_name='Correlation Summary', index=False)  # Write correlation summary to its own sheet

    # Write the third set of DataFrames (nbtrueresult) to separate sheets
    for i, df in enumerate(nbtrueresult):
        sheet_name = f'NB{i + 1} Correlation Results'  # Create a sheet name for the results
        df.to_excel(writer, sheet_name=sheet_name, index=False)  # Write the DataFrame to the specified sheet
        worksheet = writer.sheets[sheet_name]  # Access the worksheet for further formatting

        # Apply PatternFill for `Result NB# U#`
        for col_index in range(1, len(df.columns)):  # Dynamic column range
            if col_index < len(df.columns) - 1:  # Exclude the last "NB# Result" column
                column_letter = get_column_letter(col_index + 1)
                for cell in worksheet[column_letter][1:]:  # Exclude header row
                    value = cell.value
                    if value == "Passed":
                        cell.fill = PatternFill(start_color="78e08f", end_color="78e08f", fill_type="solid")
                    elif value == "For check":
                        cell.fill = PatternFill(start_color="feca57", end_color="feca57", fill_type="solid")
                    elif value == "Failed":
                        cell.fill = PatternFill(start_color="ea8685", end_color="ea8685", fill_type="solid")

        # Apply PatternFill for `NB# Result` (last column)
        result_column_letter = get_column_letter(len(df.columns))  # Last column for "NB# Result"
        for cell in worksheet[result_column_letter][1:]:  # Exclude header row
            value = cell.value
            if value == "Passed":
                cell.fill = PatternFill(start_color="78e08f", end_color="78e08f", fill_type="solid")
            elif value == "For check":
                cell.fill = PatternFill(start_color="feca57", end_color="feca57", fill_type="solid")
            elif value == "Failed":
                cell.fill = PatternFill(start_color="ea8685", end_color="ea8685", fill_type="solid")

    # Write the second set of DataFrames (nbcpkresult) to separate sheets
    for i, df in enumerate(nbcpkresult):
        sheet_name = f'NB{i + 1} CPK'  # Create a sheet name for CPK results
        df.to_excel(writer, sheet_name=sheet_name, index=False)  # Write the DataFrame to the specified sheet
        worksheet = writer.sheets[sheet_name]  # Access the worksheet for further formatting

        # Apply PatternFill for `Cpk RB Result` and `Cpk NB# Result`
        for idx, cell in enumerate(worksheet['K'][1:], start=1):  # Column K
            value = cell.value
            if value == "Good capable":
                cell.fill = PatternFill(start_color="78e08f", end_color="78e08f", fill_type="solid")
            elif value == "Not capable":
                cell.fill = PatternFill(start_color="ea8685", end_color="ea8685", fill_type="solid")
            elif value == "N/A":
                cell.fill = PatternFill(start_color="95afc0", end_color="95afc0", fill_type="solid")

        for idx, cell in enumerate(worksheet['P'][1:], start=1):  # Column P
            value = cell.value
            if value == "Good capable":
                cell.fill = PatternFill(start_color="78e08f", end_color="78e08f", fill_type="solid")
            elif value == "Not capable":
                cell.fill = PatternFill(start_color="ea8685", end_color="ea8685", fill_type="solid")
            elif value == "N/A":
                cell.fill = PatternFill(start_color="95afc0", end_color="95afc0", fill_type="solid")

    # Access the workbook and the writer's worksheets
    workbook = writer.book

    # Define a border style for the cells
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    fill_color = PatternFill(start_color='82ccdd', end_color='82ccdd', fill_type='solid')
    fill_color_info = PatternFill(start_color='74b9ff', end_color='74b9ff', fill_type='solid')
    fill_color_corr = PatternFill(None)

    # Adjust column widths, add autofilters, and apply formatting
    for sheet_name in writer.sheets:
        worksheet = workbook[sheet_name]

        # Adjust column widths for the current sheet
        for column in worksheet.columns:
            # Filter out None values and check if there are any values left
            non_empty_cells = [cell.value for cell in column if cell.value is not None]
            
            if non_empty_cells:  # Only calculate max length if the column has non-empty values
                max_length = max(len(str(cell)) for cell in non_empty_cells)
                adjusted_width = max_length + 2
            else:
                adjusted_width = 10  # Set a default width if the column is empty
        
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

            for cell in worksheet[1]:  # Access the first row
                cell.fill = fill_color  # Apply the fill to each cell in the first row

            for cell in column:
                cell.border = thin_border  # Apply border to each cell in the column

        # Add autofilter to all sheets except 'Info'
        if sheet_name != 'Info':
            worksheet.auto_filter.ref = worksheet.dimensions

        # Apply borders and formatting to specific sheets
        if sheet_name == 'Info':
            for cell in worksheet[1]:  # Access the first row
                cell.fill = fill_color_info  # Apply the fill to each cell in the first row

            for cell in worksheet[8]:  # Access the eighth row
                cell.fill = fill_color_info  # Apply the fill to each cell in the eighth row

            # Apply borders to the first table (df_product_info)
            for row in worksheet.iter_rows(min_row=1, max_row=len(df_product_info) + 1, min_col=1, max_col=len(df_product_info.columns)):
                for cell in row:
                    cell.border = thin_border  # Apply border to each cell in the first table

            # Apply borders to the second table (df_setup_info)
            for row in worksheet.iter_rows(min_row=len(df_product_info) + 3, max_row=len(df_product_info) + len(df_setup_info) + 3, min_col=1, max_col=len(df_setup_info.columns)):
                for cell in row:
                    cell.border = thin_border  # Apply border to each cell in the second table

            for cell in worksheet[7]:  # Access the eighth row
                cell.border = None  # Remove border from the eighth row

        if sheet_name == 'Correlation Summary':
            for row in worksheet.iter_rows(min_row=1, max_row=len(corrtable) + 1, min_col=1, max_col=len(corrtable.columns)):
                for cell in row:
                    cell.border = thin_border  # Apply border to each cell in the correlation summary
                    cell.fill = fill_color_corr  # Apply fill color to each cell

                    # Align all cells to the left
                    cell.alignment = Alignment(horizontal='left')

                    # Make the first column bold
                    if cell.column == 1:  # Check if it's the first column
                        cell.font = Font(bold=True)

        # Freeze the first row for all sheets except 'Info' and 'Correlation Summary'
        if sheet_name not in ['Info', 'Correlation Summary']:
            worksheet.freeze_panes = worksheet['G2']

        # Align the first column to the left for all sheets except 'Info' and 'Correlation Summary'
        if sheet_name not in ['Info', 'Correlation Summary']:
            for cell in worksheet['A'][1:]:  # Access the first column (A), excluding the header
                cell.alignment = Alignment(horizontal='left')  # Set alignment to left

# Print a message indicating that the DataFrames have been written to the specified output file
print()
print(f"Report has been written to '{output_file}'")

# Set the timer duration in seconds
timer_duration = 10  # Change this to your desired duration

# Call the thank_you function to display a message or perform an action
thank_you()

# Countdown loop to display the remaining time
for remaining in range(timer_duration, 0, -1):
    # Print the remaining time, overwriting the same line in the terminal
    print(f"The terminal will close in {remaining} seconds...", end='\r')  # Use '\r' to overwrite the line
    time.sleep(1)  # Pause execution for 1 second

# Exit the program after the countdown is complete
sys.exit()

get_ipython().system('jupyter nbconvert --to script Auto_Report_Gen-GUI.ipynb')

